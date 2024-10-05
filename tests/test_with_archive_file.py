import csv
import io
from pypdf import PdfReader
from openpyxl import load_workbook

from path_config import ARCHIVE_PATH


def test_csv(open_file_from_archive):
    content = io.StringIO(open_file_from_archive(ARCHIVE_PATH, "test_file_csv.csv").read().decode("utf-8"))
    reader = csv.reader(content)
    for index, row in enumerate(reader):
        if index == 2:
            assert row[2] == "Треугольник"


def test_pdf(open_file_from_archive):
    content = io.BytesIO(open_file_from_archive(ARCHIVE_PATH, "Python Testing with Pytest (Brian Okken).pdf").read())
    reader = PdfReader(content)
    assert "Simple, Rapid, Effective, and Scalable" in reader.get_page(1).extract_text()


def test_xlsx(open_file_from_archive):
    content = open_file_from_archive(ARCHIVE_PATH, "file_example_XLSX_50.xlsx")
    workbook = load_workbook(content)
    sheet = workbook.active
    assert sheet.cell(row=2, column=3).value == "Abril"
